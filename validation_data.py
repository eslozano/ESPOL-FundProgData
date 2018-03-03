from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from estudiante import Estudiante
import pandas as pd


#Escriba el nombre del archivo de los datos a validar
filename = "PARNN_2017II.xlsx"




OPCION_NO_VALIDA = 'dato_no_valido'
VACIO = 'dato_vacio'
FUERA_DE_RANGO = 'dato_fuera_de_rango'
NO_REDONDEADO = 'dato_no_redondeado'
TIPO_NO_NUMERICO = 'dato_no_numerico'



def fillCell(cell,color,fill_type=None):
	fill = PatternFill(fgColor=color,fill_type=fill_type)
	cell.fill = fill

def validateDataSheet(filename,columnas,colores,errorsMessages,validaciones):
	header_rows = 2
	VACIO = errorsMessages[0]
	NO_REDONDEADO = errorsMessages[-1]
	TIPO_NO_NUMERICO = errorsMessages[-2]
	try:
		wb = load_workbook(filename)
		data_sheet = wb["DATA"]
		total_students = data_sheet.max_row - header_rows
		total_columns = data_sheet.max_column 

		print("Total estudiantes a revisar: "+str(total_students))
		total_errors = 0
		for row in data_sheet.iter_rows(min_row=header_rows+1, max_col=total_columns, max_row=data_sheet.max_row):
			data = []
			for cell in row:
				data.append(cell.value)
				fillCell(cell,colores[VACIO]) #Restore cell color
			estudiante = Estudiante(data)
			errores = estudiante.validar(errorsMessages,validaciones)
			if len(errores[NO_REDONDEADO]): #Si hay valores no redondeados en las calificaciones finales
				estudiante.redondearCalifFinales(TIPO_NO_NUMERICO)
			countErrorStudent = 0
			for error, campos in errores.items():
				if error!=NO_REDONDEADO: #El error detectado es corregido por el script
					color = colores[error]
					countErrorStudent += len(campos)
					for c in campos:
						if c in columnas:
							i = columnas.index(c)
							fillCell(row[i],color,'solid')
			if countErrorStudent:
				print('{:<9}{}'.format('FALLÓ',estudiante.nombre))
				total_errors += 1
			else:
				print('{:<9}{}'.format('OK',estudiante.nombre))
		if total_errors:
			print("RESULTADOS: Se detectaron errores en las calificaciones de ", total_errors," estudiantes.")
			print("Revise en el archivo las celdas resaltadas de colores:")
			print("Naranja(celda_vacia), Azul(opcion_no_valida), Amarillo(fuera_de_rango) o Verde(dato_no_numerico).")
			print("="*100)
			print("** Los errores de dato_no_redondeado fueron corregidos por el script")
		else:
			print("="*100)
			print('¡Felicidades! Tu archivo de excel está listo para ser enviado')
			print('Por favor enviar a rabonilla@espol.edu.ec y a eslozano@espol.edu.ec')
		print("="*100)
		print("Copyright (c) 2018 eslozano")
		print('All Rights Reserved :)')
		wb.save(filename)
	except IOError as e:
		print("ERROR: No." , e.errno , e )
		if e.errno == 2:
			print("El archivo ",filename," no se encuentra en la carpeta del script.")
		if e.errno == 13:
			print("El archivo ",filename," se encuentra abierto.")

def dataAnalysis(filename):
	df = pd.read_excel("filename",skiprows =2)


#Las columnas deben estar en el orden del excel
columnas = ["nombre","matricula","genero","paralelo","cod_carrera","veces_tomadas",
			"1er_proyecto", "1er_sustent", "1er_calif_final",
			"1er_exam_tema1","1er_exam_tema2","1er_exam_tema3","1er_exam_tema4",
			"2do_proyecto", "2do_sustent", "2do_calif_final",
			"2do_exam_tema1","2do_exam_tema2","2do_exam_tema3",
			"calif_final_practica",
			"3er_proyecto","3er_calif_final",
			"3er_exam_tema1","3er_exam_tema2","3er_exam_tema3"
]
#Los colores en los que se resaltarán las celdas según los errores
colors = {
	VACIO:"ffd8bf", #naranja 
	OPCION_NO_VALIDA: '6aa2fc', #azul 
	FUERA_DE_RANGO: 'fffa00', #amarillo
	TIPO_NO_NUMERICO: '7cff94', #verde
	NO_REDONDEADO: 'ff7ff4' #rosado
}

errorsMessages= [VACIO,OPCION_NO_VALIDA,FUERA_DE_RANGO,TIPO_NO_NUMERICO,NO_REDONDEADO]

validaciones = {	
	"genero": ["F","M"],
	"veces_tomadas":[2,3],
	"calif_final": 100,
	"sustent": 1,
	"proyecto": {"1er_proyecto":20,"2do_proyecto":20,"3er_proyecto":25},
	"examen":{
		"1er_exam_": { "tema1": 20, "tema2": 32, "tema3": 45, "tema4": 10 },
		"2do_exam_": { "tema1": 35, "tema2": 55, "tema3": 10 },
		"3er_exam_": { "tema1": 40, "tema2": 55, "tema3": 10 } #el tema 2 tiene bono de 5 puntos
	}
}

validateDataSheet(filename,columnas,colors,errorsMessages, validaciones)